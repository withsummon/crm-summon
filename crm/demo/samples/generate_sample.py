"""
Generate sample financial spread files for credit analysis testing.
Output: CSV (always) and optionally XLSX (if openpyxl is installed).

Usage:
    python crm/demo/samples/generate_sample.py
    python crm/demo/samples/generate_sample.py --xlsx
"""

import csv
import os
import argparse

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

HEADERS = ["statement_type", "metric_label", "year", "amount"]

ROWS = [
    # P&L - 3 years
    ("P&L", "Pendapatan / Sales", 2022, 8500000000000),
    ("P&L", "Pendapatan / Sales", 2023, 9200000000000),
    ("P&L", "Pendapatan / Sales", 2024, 10100000000000),
    ("P&L", "Beban Pokok Penjualan", 2022, 5950000000000),
    ("P&L", "Beban Pokok Penjualan", 2023, 6440000000000),
    ("P&L", "Beban Pokok Penjualan", 2024, 7070000000000),
    ("P&L", "Laba Kotor", 2022, 2550000000000),
    ("P&L", "Laba Kotor", 2023, 2760000000000),
    ("P&L", "Laba Kotor", 2024, 3030000000000),
    ("P&L", "Beban Operasional", 2022, 1100000000000),
    ("P&L", "Beban Operasional", 2023, 1200000000000),
    ("P&L", "Beban Operasional", 2024, 1300000000000),
    ("P&L", "EBITDA", 2022, 1450000000000),
    ("P&L", "EBITDA", 2023, 1560000000000),
    ("P&L", "EBITDA", 2024, 1730000000000),
    ("P&L", "Beban Bunga", 2022, 380000000000),
    ("P&L", "Beban Bunga", 2023, 410000000000),
    ("P&L", "Beban Bunga", 2024, 450000000000),
    ("P&L", "Pajak", 2022, 220000000000),
    ("P&L", "Pajak", 2023, 250000000000),
    ("P&L", "Pajak", 2024, 280000000000),
    ("P&L", "Laba Bersih", 2022, 850000000000),
    ("P&L", "Laba Bersih", 2023, 900000000000),
    ("P&L", "Laba Bersih", 2024, 1000000000000),
    # Balance Sheet - 3 years
    ("Balance Sheet", "Kas dan Setara Kas", 2022, 1200000000000),
    ("Balance Sheet", "Kas dan Setara Kas", 2023, 1350000000000),
    ("Balance Sheet", "Kas dan Setara Kas", 2024, 1500000000000),
    ("Balance Sheet", "Piutang Usaha", 2022, 1800000000000),
    ("Balance Sheet", "Piutang Usaha", 2023, 1950000000000),
    ("Balance Sheet", "Piutang Usaha", 2024, 2100000000000),
    ("Balance Sheet", "Persediaan", 2022, 950000000000),
    ("Balance Sheet", "Persediaan", 2023, 1050000000000),
    ("Balance Sheet", "Persediaan", 2024, 1150000000000),
    ("Balance Sheet", "Aset Lancar", 2022, 3950000000000),
    ("Balance Sheet", "Aset Lancar", 2023, 4350000000000),
    ("Balance Sheet", "Aset Lancar", 2024, 4750000000000),
    ("Balance Sheet", "Total Aset", 2022, 12500000000000),
    ("Balance Sheet", "Total Aset", 2023, 13800000000000),
    ("Balance Sheet", "Total Aset", 2024, 15200000000000),
    ("Balance Sheet", "Utang Usaha", 2022, 1100000000000),
    ("Balance Sheet", "Utang Usaha", 2023, 1200000000000),
    ("Balance Sheet", "Utang Usaha", 2024, 1350000000000),
    ("Balance Sheet", "Liabilitas Jangka Pendek", 2022, 1800000000000),
    ("Balance Sheet", "Liabilitas Jangka Pendek", 2023, 2000000000000),
    ("Balance Sheet", "Liabilitas Jangka Pendek", 2024, 2200000000000),
    ("Balance Sheet", "Pinjaman Berbunga", 2022, 3500000000000),
    ("Balance Sheet", "Pinjaman Berbunga", 2023, 3800000000000),
    ("Balance Sheet", "Pinjaman Berbunga", 2024, 4200000000000),
    ("Balance Sheet", "Total Liabilitas", 2022, 6500000000000),
    ("Balance Sheet", "Total Liabilitas", 2023, 7100000000000),
    ("Balance Sheet", "Total Liabilitas", 2024, 7800000000000),
    ("Balance Sheet", "Ekuitas", 2022, 6000000000000),
    ("Balance Sheet", "Ekuitas", 2023, 6700000000000),
    ("Balance Sheet", "Ekuitas", 2024, 7400000000000),
    # Cash Flow - 3 years
    ("Cash Flow", "Arus Kas Operasi", 2022, 1200000000000),
    ("Cash Flow", "Arus Kas Operasi", 2023, 1300000000000),
    ("Cash Flow", "Arus Kas Operasi", 2024, 1450000000000),
    ("Cash Flow", "Belanja Modal", 2022, 450000000000),
    ("Cash Flow", "Belanja Modal", 2023, 500000000000),
    ("Cash Flow", "Belanja Modal", 2024, 550000000000),
    ("Cash Flow", "Arus Kas Pendanaan", 2022, 300000000000),
    ("Cash Flow", "Arus Kas Pendanaan", 2023, 350000000000),
    ("Cash Flow", "Arus Kas Pendanaan", 2024, 400000000000),
    ("Cash Flow", "Arus Kas Periode Berjalan", 2022, 1050000000000),
    ("Cash Flow", "Arus Kas Periode Berjalan", 2023, 1150000000000),
    ("Cash Flow", "Arus Kas Periode Berjalan", 2024, 1300000000000),
    ("Cash Flow", "Kas Akhir Periode", 2022, 1200000000000),
    ("Cash Flow", "Kas Akhir Periode", 2023, 1350000000000),
    ("Cash Flow", "Kas Akhir Periode", 2024, 1500000000000),
]


def write_csv(filepath):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for row in ROWS:
            writer.writerow(row)
    print(f"  CSV: {filepath} ({os.path.getsize(filepath)} bytes)")


def write_xlsx(filepath):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Spreading"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")

        # Write headers
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write rows
        for row_idx, row in enumerate(ROWS, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 4:  # amount column
                    cell.number_format = '#,##0'

        # Auto-width columns
        for col in range(1, 5):
            max_len = len(str(HEADERS[col - 1]))
            for row_idx in range(2, len(ROWS) + 2):
                cell_val = ws.cell(row=row_idx, column=col).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[chr(64 + col)].width = max_len + 4

        wb.save(filepath)
        print(f"  XLSX: {filepath} ({os.path.getsize(filepath)} bytes)")
    except ImportError:
        print("  [SKIP] XLSX: openpyxl not installed")


def main():
    parser = argparse.ArgumentParser(description="Generate sample financial spread files")
    parser.add_argument("--xlsx", action="store_true", help="Generate XLSX file (requires openpyxl)")
    parser.add_argument("--prefix", default="financial_spread_sample", help="Output filename prefix")
    args = parser.parse_args()

    print("Generating sample financial spread files...")

    csv_path = os.path.join(OUTPUT_DIR, f"{args.prefix}.csv")
    write_csv(csv_path)

    if args.xlsx:
        xlsx_path = os.path.join(OUTPUT_DIR, f"{args.prefix}.xlsx")
        write_xlsx(xlsx_path)

    print(f"\nDone. Files are in: {OUTPUT_DIR}")
    print("\nUpload CSV/XLSX via the Financial Summary tab when creating a Credit Application.")
    print("Or upload directly on the Financial Spreading tab in the analysis workspace.")
    print("\nRequired columns: statement_type, metric_label, year, amount")
    print("Statement types: P&L, Balance Sheet, Cash Flow")
    print("Must contain 3-5 years of financial data.")


if __name__ == "__main__":
    main()
