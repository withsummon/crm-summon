import os
import re
import time
import zipfile
from collections import Counter
from xml.etree import ElementTree as ET

import frappe
from frappe import _
from frappe.utils import get_datetime, nowdate

from crm.api.lead_management import capture_lead


WORKBOOK_FIELDS = [
	{"field": "sequence", "label": "Row ID"},
	{"field": "company", "label": "COMPANY"},
	{"field": "industry", "label": "Industry"},
	{"field": "phone", "label": "PHONE"},
	{"field": "name", "label": "NAME"},
	{"field": "position", "label": "POSITION"},
	{"field": "email", "label": "EMAIL"},
	{"field": "pic", "label": "PIC"},
	{"field": "status", "label": "STATUS"},
	{"field": "date", "label": "DATE"},
	{"field": "feedback", "label": "FEEDBACK"},
	{"field": "follow_up", "label": "FOLLOW UP"},
	{"field": "linkedin", "label": "LINKEDIN"},
]

FIELD_LABELS = {item["field"]: item["label"] for item in WORKBOOK_FIELDS}
FIELD_ORDER = [item["field"] for item in WORKBOOK_FIELDS]

HEADER_ALIASES = {
	"company": {"company", "organization", "perusahaan", "nama entitas"},
	"industry": {"industry", "industri", "sektor"},
	"phone": {"phone", "mobile", "mobile_no", "whatsapp", "wa", "telp", "no hp"},
	"name": {"name", "lead_name", "contact_name", "contact", "nama"},
	"position": {"position", "designation", "job_title", "title", "jabatan"},
	"email": {"email", "email_address", "email_id", "surel"},
	"pic": {"pic", "owner", "relationship_manager", "rm"},
	"status": {"status", "lead_status"},
	"date": {"date", "created_at", "tanggal"},
	"feedback": {"feedback", "notes", "catatan"},
	"follow_up": {"follow up", "follow_up", "next step", "next_step"},
	"linkedin": {"linkedin", "linkedin_url", "website", "url"},
	"sequence": {"no", "number", "sequence", "row id"},
}

STATUS_MAP = {
	"open": "New",
	"new": "New",
	"contacted": "Contacted",
	"qualified": "Qualified",
	"converted": "Converted",
	"closed": "Converted",
	"won": "Converted",
	"lost": "Unqualified",
	"unqualified": "Unqualified",
	"junk": "Junk",
	"do not contact": "Junk",
}

XML_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

DEFAULT_IMPORT_OPTIONS = {
	"default_source": "BNI Lead Workbook",
	"default_channel": "Excel Import",
	"create_follow_up_tasks": 1,
	"create_notes": 1,
}


def _doctype_ready(doctype):
	try:
		return frappe.db.table_exists(doctype)
	except Exception:
		return False


def _load_import_options(options=None, **kwargs):
	parsed = {}
	if isinstance(options, str):
		parsed = frappe.parse_json(options) or {}
	elif isinstance(options, dict):
		parsed = dict(options)

	merged = {**DEFAULT_IMPORT_OPTIONS, **parsed}
	for key, value in kwargs.items():
		if value is not None:
			merged[key] = value
	merged["create_follow_up_tasks"] = int(merged.get("create_follow_up_tasks") or 0)
	merged["create_notes"] = int(merged.get("create_notes") or 0)
	merged["default_source"] = (merged.get("default_source") or DEFAULT_IMPORT_OPTIONS["default_source"]).strip()
	merged["default_channel"] = (merged.get("default_channel") or DEFAULT_IMPORT_OPTIONS["default_channel"]).strip()
	return merged


def _normalize_header(value):
	value = re.sub(r"[_\-]+", " ", str(value or "").strip().lower())
	return re.sub(r"\s+", " ", value).strip()


def _normalize_text(value):
	value = str(value or "").strip()
	if not value or value.lower() in {"none", "null", "nan"}:
		return None
	return value


def _normalize_phone(value):
	value = _normalize_text(value)
	if not value:
		return None
	value = value.replace("https://", "").replace("http://", "")
	if "wa.me/" in value:
		value = value.split("wa.me/")[-1]
	value = value.replace("whatsapp:", "")
	value = re.sub(r"[^\d+]", "", value)
	if value.startswith("62") and not value.startswith("+62"):
		return f"+{value}"
	if value.startswith("0"):
		return f"+62{value[1:]}"
	return value


def _normalize_email(value):
	value = _normalize_text(value)
	if not value or "@" not in value:
		return None
	return value.lower()


def _normalize_link(value):
	value = _normalize_text(value)
	if not value:
		return None
	if value.startswith("www."):
		return f"https://{value}"
	if "://" not in value:
		return f"https://{value}"
	return value


def _normalize_status(value):
	value = _normalize_text(value)
	if not value:
		return "New"
	return STATUS_MAP.get(value.lower(), "New")


def _safe_datetime(value):
	value = _normalize_text(value)
	if not value:
		return None
	for candidate in (value, value.replace(".", "-"), value.replace("/", "-")):
		try:
			return get_datetime(candidate)
		except Exception:
			continue
	return None


def _safe_date_string(value):
	date_value = _safe_datetime(value)
	return date_value.strftime("%Y-%m-%d %H:%M:%S") if date_value else None


def _resolve_user_from_pic(pic):
	pic = _normalize_text(pic)
	if not pic:
		return None

	for filters in (
		{"name": pic},
		{"email": pic},
		{"full_name": pic},
	):
		user = frappe.db.get_value("User", filters, "name")
		if user:
			return user
	return None


def _normalize_industry(value):
	value = _normalize_text(value)
	if not value or not _doctype_ready("CRM Industry"):
		return None
	if frappe.db.exists("CRM Industry", value):
		return value
	return None


def _column_to_index(ref):
	column = "".join(ch for ch in ref if ch.isalpha()).upper()
	value = 0
	for ch in column:
		value = value * 26 + ord(ch) - 64
	return value - 1


def _iter_sheet_rows_with_openpyxl(file_path):
	import openpyxl

	workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	worksheet = workbook.active
	rows = []
	max_columns = worksheet.max_column or len(FIELD_ORDER)
	for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
		values = list(row[:max_columns])
		if len(values) < max_columns:
			values.extend([None] * (max_columns - len(values)))
		rows.append({"row_number": row_idx, "values": values})
	workbook.close()
	return rows


def _iter_sheet_rows_from_xml(file_path):
	with zipfile.ZipFile(file_path) as archive:
		root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
		rows = []
		max_columns = 0
		for row in root.findall(".//a:sheetData/a:row", XML_NS):
			values = {}
			for cell in row.findall("a:c", XML_NS):
				ref = cell.attrib.get("r", "")
				index = _column_to_index(ref) if ref else len(values)
				cell_type = cell.attrib.get("t")
				if cell_type == "inlineStr":
					text = "".join(t.text or "" for t in cell.iterfind(".//a:t", XML_NS))
				else:
					value_node = cell.find("a:v", XML_NS)
					text = value_node.text if value_node is not None else ""
				values[index] = text
				max_columns = max(max_columns, index + 1)
			rows.append({"row_number": int(row.attrib.get("r") or len(rows) + 1), "cells": values})

	if not max_columns:
		return []

	normalized = []
	for row in rows:
		out = [None] * max_columns
		for index, value in row.get("cells", {}).items():
			if index < max_columns:
				out[index] = value
		normalized.append({"row_number": row["row_number"], "values": out})
	return normalized


def _load_sheet_rows(file_path):
	try:
		return _iter_sheet_rows_with_openpyxl(file_path)
	except Exception:
		return _iter_sheet_rows_from_xml(file_path)


def _resolve_file_path(file_url):
	file_url = (file_url or "").strip()
	if not file_url:
		frappe.throw(_("File URL is required"))
	file_path = frappe.get_site_path(file_url.lstrip("/"))
	if not file_path or not os.path.exists(file_path):
		frappe.throw(_("File not found"))
	return file_path


def _bundled_workbook_path():
	current_dir = os.path.dirname(os.path.abspath(__file__))
	return os.path.abspath(os.path.join(current_dir, "..", "..", "Lead Gen Data.xlsx"))


def _detect_columns(headers):
	mapping = {}
	unmapped = []
	for index, header in enumerate(headers):
		header_normalized = _normalize_header(header)
		if not header_normalized:
			continue
		matched_field = None
		for fieldname, aliases in HEADER_ALIASES.items():
			if header_normalized in aliases:
				matched_field = fieldname
				break
		if not matched_field:
			unmapped.append({"index": index, "header": str(header or "")})
			continue
		if matched_field not in mapping:
			mapping[matched_field] = {"index": index, "header": str(header or ""), "field": matched_field}
	return mapping, unmapped


def _row_to_raw_dict(values, column_map):
	row = {}
	for fieldname in FIELD_ORDER:
		meta = column_map.get(fieldname)
		value = values[meta["index"]] if meta and meta["index"] < len(values) else None
		row[fieldname] = _normalize_text(value)
	return row


def _row_to_normalized(raw_row, row_number, warnings):
	normalized = dict(raw_row)
	normalized["row_number"] = row_number
	normalized["phone"] = _normalize_phone(raw_row.get("phone"))
	normalized["email"] = _normalize_email(raw_row.get("email"))
	normalized["linkedin"] = _normalize_link(raw_row.get("linkedin"))
	normalized["status"] = _normalize_status(raw_row.get("status"))
	normalized["name"] = _normalize_text(raw_row.get("name"))
	normalized["company"] = _normalize_text(raw_row.get("company"))
	normalized["feedback"] = _normalize_text(raw_row.get("feedback"))
	normalized["follow_up"] = _normalize_text(raw_row.get("follow_up"))
	normalized["pic"] = _normalize_text(raw_row.get("pic"))
	normalized["position"] = _normalize_text(raw_row.get("position"))
	normalized["industry"] = _normalize_industry(raw_row.get("industry"))
	normalized["date"] = _safe_date_string(raw_row.get("date"))
	normalized["owner"] = _resolve_user_from_pic(normalized.get("pic"))
	normalized["is_empty"] = not any(_normalize_text(raw_row.get(fieldname)) for fieldname in FIELD_ORDER)
	normalized["is_importable"] = bool(normalized.get("company") or normalized.get("name"))

	if raw_row.get("phone") and not normalized.get("phone"):
		warnings.append({"row": row_number, "field": "phone", "message": _("Phone number could not be normalized.")})
	if raw_row.get("email") and not normalized.get("email"):
		warnings.append({"row": row_number, "field": "email", "message": _("Email address is invalid and will be skipped.")})
	if raw_row.get("pic") and not normalized.get("owner"):
		warnings.append({"row": row_number, "field": "pic", "message": _("PIC does not match an existing user; auto-assignment will be used.")})
	if raw_row.get("status") and normalized.get("status") == "New" and _normalize_text(raw_row.get("status")).lower() not in STATUS_MAP:
		warnings.append({"row": row_number, "field": "status", "message": _("Status was not recognized and defaulted to New.")})
	if raw_row.get("industry") and not normalized.get("industry"):
		warnings.append({"row": row_number, "field": "industry", "message": _("Industry does not match CRM Industry and will be skipped.")})
	if not normalized["is_importable"] and not normalized["is_empty"]:
		warnings.append({"row": row_number, "field": "company", "message": _("Row is missing both company and contact name and will be skipped.")})

	return normalized


def _summarize_warnings(warnings):
	counter = Counter(item["message"] for item in warnings)
	return [{"message": message, "count": count} for message, count in counter.items()]


def _build_workbook_payload_from_path(file_path):
	rows = _load_sheet_rows(file_path)
	if not rows:
		return {
			"headers": [FIELD_LABELS[fieldname] for fieldname in FIELD_ORDER],
			"rows": [],
			"row_objects": [],
			"total_rows": 0,
			"normalized_row_count": 0,
			"empty_row_count": 0,
			"column_mapping": [],
			"warnings": [{"message": _("Workbook is empty."), "count": 1}],
		}

	header_values = rows[0]["values"]
	column_map, unmapped = _detect_columns(header_values)
	raw_headers = [str(value or "") for value in header_values]
	row_objects = []
	row_warnings = []
	empty_row_count = 0
	normalized_row_count = 0

	for row in rows[1:]:
		raw_row = _row_to_raw_dict(row["values"], column_map)
		normalized_row = _row_to_normalized(raw_row, row["row_number"], row_warnings)
		if normalized_row["is_empty"]:
			empty_row_count += 1
			continue
		if normalized_row["is_importable"]:
			normalized_row_count += 1
		row_objects.append(normalized_row)

	column_mapping = []
	for fieldname in FIELD_ORDER:
		meta = column_map.get(fieldname)
		column_mapping.append(
			{
				"field": fieldname,
				"label": FIELD_LABELS[fieldname],
				"header": meta["header"] if meta else "",
				"index": meta["index"] if meta else None,
				"mapped": bool(meta),
			}
		)

	workbook_warnings = []
	if raw_headers and not _normalize_text(raw_headers[0]):
		workbook_warnings.append({"message": _("Workbook contains a leading blank helper column that will be ignored."), "count": 1})
	if unmapped:
		workbook_warnings.append({"message": _("Workbook has unmapped columns that will be ignored."), "count": len(unmapped)})
	missing_required = [FIELD_LABELS[fieldname] for fieldname in ("company", "name", "phone", "email", "status") if fieldname not in column_map]
	if missing_required:
		workbook_warnings.append(
			{"message": _("Expected workbook columns are missing: {0}.").format(", ".join(missing_required)), "count": len(missing_required)}
		)

	workbook_warnings.extend(_summarize_warnings(row_warnings))

	return {
		"headers": [FIELD_LABELS[fieldname] for fieldname in FIELD_ORDER],
		"raw_headers": raw_headers,
		"rows": [[row.get(fieldname) or "" for fieldname in FIELD_ORDER] for row in row_objects],
		"row_objects": row_objects,
		"total_rows": len(row_objects),
		"normalized_row_count": normalized_row_count,
		"empty_row_count": empty_row_count,
		"column_mapping": column_mapping,
		"warnings": workbook_warnings,
	}


def _build_workbook_payload(file_url):
	return _build_workbook_payload_from_path(_resolve_file_path(file_url))


def _ensure_note_for_lead(lead_name, row):
	if not _doctype_ready("FCRM Note"):
		return None
	content_bits = []
	if row.get("feedback"):
		content_bits.append(f"Feedback: {row['feedback']}")
	if row.get("follow_up"):
		content_bits.append(f"Follow Up: {row['follow_up']}")
	if row.get("pic"):
		content_bits.append(f"PIC: {row['pic']}")
	if row.get("linkedin"):
		content_bits.append(f"LinkedIn/Website: {row['linkedin']}")
	if not content_bits:
		return None

	return frappe.get_doc(
		{
			"doctype": "FCRM Note",
			"title": f"Lead Gen Workbook Context #{row['row_number']}",
			"content": "\n".join(content_bits),
			"reference_doctype": "CRM Lead",
			"reference_docname": lead_name,
		}
	).insert(ignore_permissions=True).name


def _ensure_follow_up_task(lead_name, row, assigned_to=None):
	if not _doctype_ready("CRM Task"):
		return None
	if not any(row.get(fieldname) for fieldname in ("follow_up", "feedback", "date")):
		return None

	description = row.get("follow_up") or row.get("feedback") or _("Imported workbook follow-up")
	title_target = row.get("company") or row.get("name") or lead_name
	due_date = _safe_datetime(row.get("follow_up")) or _safe_datetime(row.get("date"))
	if due_date is None:
		due_date = get_datetime(nowdate())

	return frappe.get_doc(
		{
			"doctype": "CRM Task",
			"title": f"Lead Gen Follow-up - {title_target}",
			"description": description,
			"priority": "Medium",
			"status": "Todo",
			"assigned_to": assigned_to or frappe.session.user,
			"due_date": due_date.strftime("%Y-%m-%d %H:%M:%S"),
			"reference_doctype": "CRM Lead",
			"reference_docname": lead_name,
		}
	).insert(ignore_permissions=True).name


def _import_rows(row_objects, options, limit=None, publish_progress=False):
	created = 0
	skipped = 0
	duplicate_count = 0
	errors = []
	notes_created = 0
	tasks_created = 0
	import_warnings = []
	processed = 0
	lead_names = []
	note_names = []
	task_names = []

	for row in row_objects:
		if limit and processed >= limit:
			break
		if not row.get("is_importable"):
			continue
		processed += 1

		try:
			payload = {
				"lead_name": row.get("name") or row.get("company"),
				"first_name": (row.get("name") or row.get("company") or "").split()[0] if (row.get("name") or row.get("company")) else "",
				"last_name": " ".join((row.get("name") or "").split()[1:]) if row.get("name") else "",
				"email": row.get("email"),
				"mobile_no": row.get("phone"),
				"organization": row.get("company"),
				"job_title": row.get("position"),
				"industry": row.get("industry"),
				"status": row.get("status") or "New",
				"source": options["default_source"],
				"capture_channel": options["default_channel"],
				"website": row.get("linkedin"),
				"lead_owner": row.get("owner"),
				"referrer_type": "Other" if row.get("pic") else None,
				"referrer": row.get("pic"),
			}

			result = capture_lead(payload, channel=options["default_channel"])
			if not result.get("created"):
				skipped += 1
				if result.get("duplicate"):
					duplicate_count += 1
				continue

			lead_name = result.get("lead")
			created += 1
			lead_names.append(lead_name)

			note_name = _ensure_note_for_lead(lead_name, row) if options["create_notes"] else None
			if note_name:
				notes_created += 1
				note_names.append(note_name)
			task_name = _ensure_follow_up_task(lead_name, row, row.get("owner")) if options["create_follow_up_tasks"] else None
			if task_name:
				tasks_created += 1
				task_names.append(task_name)

			if publish_progress:
				frappe.publish_realtime(
					"lead_scraped",
					{"lead_name": payload["lead_name"], "company": payload.get("organization"), "count": created},
					user=frappe.session.user,
				)
				time.sleep(0.25)
		except Exception as exc:
			errors.append({"row": row.get("row_number"), "error": str(exc)})
			frappe.log_error(frappe.get_traceback(), _("Lead Workbook Import Error"))
			if len(errors) >= 20:
				break

	if duplicate_count:
		import_warnings.append({"message": _("Duplicate leads were skipped during import."), "count": duplicate_count})

	frappe.db.commit()
	frappe.clear_cache(doctype="CRM Lead")
	return {
		"created": created,
		"skipped": skipped,
		"errors": errors[:10],
		"warnings": import_warnings,
		"duplicates": duplicate_count,
		"notes_created": notes_created,
		"follow_ups_created": tasks_created,
		"total": processed,
		"lead_names": lead_names,
		"note_names": note_names,
		"task_names": task_names,
	}


@frappe.whitelist()
def import_leads_from_excel(
	file_url: str,
	options=None,
	default_source: str | None = None,
	default_channel: str | None = None,
	create_follow_up_tasks: int | None = None,
	create_notes: int | None = None,
):
	payload = _build_workbook_payload(file_url)
	import_options = _load_import_options(
		options,
		default_source=default_source,
		default_channel=default_channel,
		create_follow_up_tasks=create_follow_up_tasks,
		create_notes=create_notes,
	)
	result = _import_rows(payload["row_objects"], import_options)
	result["warnings"] = payload["warnings"] + result["warnings"]
	result["normalized_rows"] = payload["normalized_row_count"]
	result["mapping"] = payload["column_mapping"]
	return result


@frappe.whitelist()
def get_excel_preview(file_url: str):
	payload = _build_workbook_payload(file_url)
	return {
		"headers": payload["headers"],
		"rows": payload["rows"][:10],
		"total_rows": payload["total_rows"],
		"normalized_row_count": payload["normalized_row_count"],
		"empty_row_count": payload["empty_row_count"],
		"column_mapping": payload["column_mapping"],
		"warnings": payload["warnings"],
		"importable_fields": [FIELD_LABELS[fieldname] for fieldname in FIELD_ORDER if fieldname != "sequence"],
		"default_options": DEFAULT_IMPORT_OPTIONS,
	}


@frappe.whitelist()
def mock_lead_scraping(prompt: str | None = None):
	file_path = _bundled_workbook_path()
	if not os.path.exists(file_path):
		frappe.throw(_("Excel file not found at {0}").format(file_path))

	payload = _build_workbook_payload_from_path(file_path)
	row_objects = payload["row_objects"]
	limit = 15

	if prompt:
		match_count = re.search(r"\b(\d+)\b", prompt)
		if match_count:
			limit = min(int(match_count.group(1)), 50)

		filter_terms = []
		cleaned = prompt.lower()
		for token in ["carikan", "saya", "tolong", "find", "me", "get", "lead", "data", "orang", "buatkan", "cariin", "dari"]:
			cleaned = cleaned.replace(token, " ")
		filter_terms = [term.strip() for term in cleaned.split() if term.strip() and not term.isdigit()]
		if filter_terms:
			filtered_rows = []
			for row in row_objects:
				haystack = " ".join(str(row.get(fieldname) or "").lower() for fieldname in FIELD_ORDER)
				if any(term in haystack for term in filter_terms):
					filtered_rows.append(row)
			if filtered_rows:
				row_objects = filtered_rows

	frappe.publish_realtime("lead_gen_start", {"total": min(limit, len(row_objects))}, user=frappe.session.user)
	result = _import_rows(row_objects, DEFAULT_IMPORT_OPTIONS, limit=limit, publish_progress=True)
	frappe.publish_realtime("lead_gen_complete", {"created": result["created"]}, user=frappe.session.user)
	frappe.clear_cache(doctype="CRM Lead")
	return {
		"message": _("Successfully imported {0} new leads. Skipped {1}.").format(result["created"], result["skipped"]),
		"created": result["created"],
		"warnings": payload["warnings"] + result["warnings"],
	}
